from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Iterable, Sequence
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


DEFAULT_TIMEZONE = "America/Costa_Rica"
KG_TO_LB = Decimal("2.2046226218")


@dataclass(frozen=True)
class DimensionGroup:
    source_type: str
    trailer: str
    barcode: str
    pallets: int
    boxes: int
    weight_kg: Decimal

    @property
    def weight_lb(self) -> int:
        return int((self.weight_kg * KG_TO_LB).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


@dataclass(frozen=True)
class OutputRow:
    carrier: str
    trailer: str
    barcode: str
    length: int
    width: int
    height: int
    weight_lb: int
    boxes: int
    scanned_time: str
    rack: str = ""
    hazmat: str = ""
    time_in: str = ""
    time_out: str = ""


class DimensionConversionError(ValueError):
    pass


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("#", " number ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _canonical_number(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return Decimal(text)
    except Exception:
        return None


def _canonical_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def _numeric_code_key(value: str) -> tuple[int, int | str, str]:
    text = str(value).strip()
    if text.isdigit():
        return (0, int(text), text)
    return (1, text, text)


def _read_xlsx(file_bytes: bytes) -> list[tuple[str, list[list[Any]]]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheets: list[tuple[str, list[list[Any]]]] = []
    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        sheets.append((worksheet.title, rows))
    workbook.close()
    return sheets


def _read_xls(file_bytes: bytes) -> list[tuple[str, list[list[Any]]]]:
    try:
        import xlrd
    except ImportError as error:
        raise DimensionConversionError(
            "Falta la libreria xlrd para leer archivos .xls. Revisa requirements.txt."
        ) from error

    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheets: list[tuple[str, list[list[Any]]]] = []
    for worksheet in workbook.sheets():
        rows = [worksheet.row_values(row_index) for row_index in range(worksheet.nrows)]
        sheets.append((worksheet.name, rows))
    return sheets


def _read_sheets(file_bytes: bytes, filename: str) -> list[tuple[str, list[list[Any]]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls":
        return _read_xls(file_bytes)
    if suffix == ".xlsx":
        return _read_xlsx(file_bytes)
    raise DimensionConversionError("El archivo debe ser .xlsx o .xls.")


def _find_header(rows: Sequence[Sequence[Any]]) -> tuple[int, dict[str, int], str] | None:
    aliases = {
        "tracking": {"tracking", "airwaybill", "air waybill", "barcode"},
        "weight": {"weight", "weight per dn", "actual weight"},
        "boxes": {"box qty", "box qty per dn", "boxes", "box"},
        "pallets": {"pallets", "pallet"},
        "trailer": {"container", "truck", "truck number", "truck no", "trailer", "trailer number"},
    }

    for row_index, row in enumerate(rows[:100]):
        normalized = [_normalize_header(value) for value in row]
        column_map: dict[str, int] = {}
        for field, names in aliases.items():
            for column_index, header in enumerate(normalized):
                if header in names:
                    column_map[field] = column_index
                    break
        required = {"tracking", "weight", "boxes", "pallets"}
        if required.issubset(column_map):
            source_type = "VRP" if "trailer" in column_map and normalized[column_map["trailer"]] == "container" else "BEX"
            return row_index, column_map, source_type
    return None


def _cell(rows: Sequence[Sequence[Any]], row: int, col: int) -> Any:
    if row < 0 or row >= len(rows):
        return None
    if col < 0 or col >= len(rows[row]):
        return None
    return rows[row][col]


def _find_bex_trailer(sheets: Sequence[tuple[str, Sequence[Sequence[Any]]]], filename: str) -> str:
    filename_match = re.search(r"\bAE\d{4,}\b", filename, flags=re.IGNORECASE)
    fallback = filename_match.group(0).upper() if filename_match else ""

    for _, rows in sheets:
        for row_index, row in enumerate(rows):
            for column_index, value in enumerate(row):
                text = _normalize_header(value)
                raw = _canonical_code(value)
                direct = re.fullmatch(r"AE\d{4,}", raw, flags=re.IGNORECASE)
                if direct:
                    return raw.upper()
                if text in {"truck", "truck number", "truck no", "trailer", "trailer number"}:
                    for row_offset, col_offset in ((0, 1), (0, 2), (1, 0), (1, 1), (-1, 1)):
                        candidate = _canonical_code(_cell(rows, row_index + row_offset, column_index + col_offset))
                        if re.fullmatch(r"[A-Z]{1,4}\d{3,}", candidate, flags=re.IGNORECASE):
                            return candidate.upper()
    return fallback


def _parse_groups(file_bytes: bytes, filename: str) -> list[DimensionGroup]:
    sheets = _read_sheets(file_bytes, filename)
    bex_trailer = _find_bex_trailer(sheets, filename)
    grouped: dict[tuple[str, str, str], dict[str, Decimal]] = {}
    detected_types: set[str] = set()

    for _, rows in sheets:
        header = _find_header(rows)
        if header is None:
            continue
        header_row, columns, source_type = header
        detected_types.add(source_type)

        for row in rows[header_row + 1:]:
            barcode = _canonical_code(row[columns["tracking"]] if columns["tracking"] < len(row) else None)
            if not barcode or not re.search(r"\d", barcode):
                continue

            weight = _canonical_number(row[columns["weight"]] if columns["weight"] < len(row) else None)
            boxes = _canonical_number(row[columns["boxes"]] if columns["boxes"] < len(row) else None)
            pallets = _canonical_number(row[columns["pallets"]] if columns["pallets"] < len(row) else None)

            if weight is None and boxes is None and pallets is None:
                continue

            if source_type == "VRP":
                trailer = _canonical_code(row[columns["trailer"]] if columns["trailer"] < len(row) else None)
            else:
                trailer = bex_trailer

            if not trailer:
                raise DimensionConversionError(
                    f"No se pudo identificar el trailer del tracking {barcode}."
                )

            key = (source_type, trailer, barcode)
            bucket = grouped.setdefault(
                key,
                {"weight": Decimal("0"), "boxes": Decimal("0"), "pallets": Decimal("0")},
            )
            bucket["weight"] += weight or Decimal("0")
            bucket["boxes"] += boxes or Decimal("0")
            if pallets is not None and pallets > 0:
                bucket["pallets"] += pallets

    if not grouped:
        raise DimensionConversionError(
            "No se encontraron encabezados validos de VRP o BEX en el archivo."
        )

    records: list[DimensionGroup] = []
    for (source_type, trailer, barcode), totals in grouped.items():
        pallet_count = int(totals["pallets"].quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        box_count = int(totals["boxes"].quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        if pallet_count <= 0:
            pallet_count = 1
        if box_count < 0:
            box_count = 0
        records.append(
            DimensionGroup(
                source_type=source_type,
                trailer=trailer,
                barcode=barcode,
                pallets=pallet_count,
                boxes=box_count,
                weight_kg=totals["weight"],
            )
        )

    records.sort(
        key=lambda record: (
            _numeric_code_key(record.barcode),
            _numeric_code_key(record.trailer),
            record.source_type,
        )
    )
    return records


def _distribute_integer(total: int, parts: int) -> list[int]:
    if parts <= 0:
        raise ValueError("parts must be positive")
    base, remainder = divmod(total, parts)
    return [base + (1 if index < remainder else 0) for index in range(parts)]


def _format_scanned_time(now: datetime | None = None) -> str:
    if now is None:
        now = datetime.now(ZoneInfo(DEFAULT_TIMEZONE))
    elif now.tzinfo is None:
        now = now.replace(tzinfo=ZoneInfo(DEFAULT_TIMEZONE))
    months = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")
    return f"{months[now.month - 1]} {now.day:02d} {now.year}"


def build_output_rows(groups: Sequence[DimensionGroup], now: datetime | None = None) -> list[OutputRow]:
    scanned_time = _format_scanned_time(now)
    rows: list[OutputRow] = []
    ordered_groups = sorted(
        groups,
        key=lambda group: (
            _numeric_code_key(group.barcode),
            _numeric_code_key(group.trailer),
            group.source_type,
        ),
    )
    for group in ordered_groups:
        box_distribution = _distribute_integer(group.boxes, group.pallets)
        weight_distribution = _distribute_integer(group.weight_lb, group.pallets)
        for boxes, weight in zip(box_distribution, weight_distribution):
            rows.append(
                OutputRow(
                    carrier=group.source_type,
                    trailer=group.trailer,
                    barcode=group.barcode,
                    length=48,
                    width=40,
                    height=40,
                    weight_lb=weight,
                    boxes=boxes,
                    scanned_time=scanned_time,
                )
            )
    return rows


def _copy_row_style(worksheet, source_row: int, target_row: int, max_column: int = 13) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.protection = copy(source.protection)


def create_output_workbook(
    output_rows: Sequence[OutputRow],
    template_path: str | Path,
) -> bytes:
    workbook = load_workbook(str(template_path))
    worksheet = workbook.active

    # The bundled template already contains the desired headers and body formatting.
    for row in range(2, max(worksheet.max_row, len(output_rows) + 1) + 1):
        if row > worksheet.max_row:
            _copy_row_style(worksheet, 2, row)
        for column in range(1, 14):
            worksheet.cell(row, column).value = None

    for row_index, record in enumerate(output_rows, start=2):
        if row_index > worksheet.max_row:
            _copy_row_style(worksheet, 2, row_index)
        barcode_value: Any = int(record.barcode) if record.barcode.isdigit() else record.barcode
        values = [
            record.carrier,
            record.trailer,
            barcode_value,
            record.length,
            record.width,
            record.height,
            record.weight_lb,
            record.boxes,
            record.rack or None,
            record.hazmat or None,
            record.scanned_time,
            record.time_in or None,
            record.time_out or None,
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row_index, column).value = value

    # Export a plain value-only range: no Excel tables and no filter dropdowns.
    for table_name in list(worksheet.tables.keys()):
        del worksheet.tables[table_name]
    worksheet.auto_filter.ref = None
    worksheet.freeze_panes = None

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def convert_dimension_file(
    file_bytes: bytes,
    filename: str,
    template_path: str | Path,
    now: datetime | None = None,
) -> tuple[bytes, list[DimensionGroup], list[OutputRow]]:
    groups = _parse_groups(file_bytes, filename)
    rows = build_output_rows(groups, now=now)
    workbook_bytes = create_output_workbook(rows, template_path)
    return workbook_bytes, groups, rows
