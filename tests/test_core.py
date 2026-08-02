from datetime import datetime
from decimal import Decimal

from bol_generator import BolRecord, apply_destination_rule
from dimension_converter import DimensionGroup, build_output_rows


def test_destination_rules():
    mexico = BolRecord("1", "P", "H", ("Cliente", "Mexico"), 1, "BEX")
    assert apply_destination_rule(mexico).destination_rule == "Mexico -> Laredo"

    puerto_rico = BolRecord("2", "P", "H", ("Cliente", "Puerto Rico"), 1, "BEX")
    assert apply_destination_rule(puerto_rico).ship_to_lines == ("Cliente", "Puerto Rico")


def test_integer_distribution_preserves_totals():
    group = DimensionGroup("BEX", "AE53304", "98161301", 2, 5, Decimal("48.56"))
    rows = build_output_rows([group], now=datetime(2026, 8, 1))
    assert len(rows) == 2
    assert sum(row.boxes for row in rows) == 5
    assert sum(row.weight_lb for row in rows) == 107
    assert all((row.length, row.width, row.height) == (48, 40, 40) for row in rows)


def test_output_row_keeps_editable_fields():
    from dimension_converter import OutputRow

    row = OutputRow(
        "BEX", "AE0001", "123", 48, 40, 40, 100, 2, "Aug 01 2026",
        rack="R1", hazmat="N", time_in="10:00 AM", time_out="10:50 AM",
    )
    assert row.rack == "R1"
    assert row.time_out == "10:50 AM"


def test_waybills_are_sorted_numerically():
    groups = [
        DimensionGroup("BEX", "AE1", "98161378", 1, 1, Decimal("1")),
        DimensionGroup("BEX", "AE1", "98158040", 1, 1, Decimal("1")),
        DimensionGroup("BEX", "AE1", "98161301", 1, 1, Decimal("1")),
    ]
    rows = build_output_rows(groups, now=datetime(2026, 8, 1))
    assert [row.barcode for row in rows] == ["98158040", "98161301", "98161378"]


def test_export_has_values_without_filters_or_formulas(tmp_path):
    from io import BytesIO
    from pathlib import Path

    from openpyxl import load_workbook

    from dimension_converter import OutputRow, create_output_workbook

    template = Path(__file__).resolve().parents[1] / "assets" / "DIMENSIONES_TEMPLATE.xlsx"
    output = create_output_workbook(
        [OutputRow("VRP", "25333", "3044657", 48, 40, 40, 107, 5, "Aug 01 2026")],
        template,
    )
    workbook = load_workbook(BytesIO(output), data_only=False)
    worksheet = workbook.active

    assert worksheet.auto_filter.ref is None
    assert not worksheet.tables
    assert worksheet.freeze_panes is None
    assert worksheet["C2"].value == 3044657
    assert worksheet["G2"].value == 107
    assert not [
        cell.coordinate
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    workbook.close()
